"""
Genera 10 archivos Excel individuales con el formato oficial SENA GFPI-F-147
(Bitácora de seguimiento Etapa productiva), usando el contenido de las
bitácoras markdown del proyecto.

Replica exactamente la estructura, estilos, fuentes, anchos de columna,
alturas de fila y celdas combinadas del formato original.
"""
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BITACORAS_DIR = os.path.dirname(os.path.abspath(__file__))

# ─── Datos generales del aprendiz y empresa ───────────────────────────────────
DATOS = {
    "empresa": "CARPAS VENTO SAS",
    "nit": 900672365,
    "jefe_nombre": "Camila Gomez",
    "jefe_telefono": 3112513220,
    "jefe_correo": "camila@ventotents.com",
    "aprendiz_nombre": "Jhon Anderson Moreno Rodriguez",
    "aprendiz_doc": 1078371141,
    "aprendiz_tel": 3204143661,
    "aprendiz_correo": "anderson960616@gmail.com",
    "ficha": 2879689,
    "programa": "ANALISIS Y DESARROLLO DE SOFTWARE",
    "regional": "REGIONAL DISTRITO CAPITAL",
    "centro": "Centro de servicios financieros",
}

# ─── Contenido de cada bitácora ───────────────────────────────────────────────

BITACORAS_DATA = [
    {
        "numero": 1,
        "periodo": "04/10/2025 - 17/10/2025",
        "fecha_entrega": datetime(2025, 10, 17),
        "actividades": [
            {
                "descripcion": (
                    "SEMANA 1-2: INVESTIGACIÓN INICIAL Y ANÁLISIS DEL DOMINIO"
                    "                    "
                    "Actividades realizadas:"
                    "                                                                                      "
                    "1. Investigación de antecedentes: Se revisaron sistemas de gestión de inventario "
                    "existentes en el mercado, incluyendo soluciones SaaS como Sortly, inFlow Inventory "
                    "y Zoho Inventory, identificando fortalezas y limitaciones para el contexto de "
                    "alquiler de carpas y eventos."
                    "                                                                                "
                    "2. Análisis del dominio del negocio: Se realizaron reuniones con el negocio de "
                    "alquiler de carpas para comprender el flujo operativo completo: recepción de "
                    "pedidos, cotización, despacho, montaje, desmontaje y retorno al inventario."
                    "                                                                                                                   "
                    "3. Identificación del problema: Se documentó que el negocio actualmente maneja "
                    "el inventario de manera manual (hojas de cálculo y registros en papel), lo que "
                    "genera errores en la disponibilidad de productos, pérdida de elementos y "
                    "dificultad para cotizar eventos."
                    "                                                "
                    "4. Revisión de literatura: Se consultaron artículos académicos sobre sistemas de "
                    "gestión de inventario, metodologías ágiles de desarrollo de software y "
                    "arquitecturas de aplicaciones web modernas."
                ),
                "fecha_inicio": datetime(2025, 10, 4),
                "fecha_fin": datetime(2025, 10, 17),
                "evidencia": (
                    "Documento de antecedentes con 5 sistemas analizados y sus "
                    "características principales"
                    "                                                      "
                    "- Mapa del flujo operativo del negocio de alquiler de carpas"
                    "                                                    "
                    "- Identificación de los problemas que el sistema debe resolver: "
                    "control de stock, trazabilidad por series/lotes, cotizaciones y "
                    "seguimiento de operaciones"
                ),
                "observaciones": (
                    "DIFICULTAD PRINCIPAL:     "
                    "La información sobre sistemas especializados "
                    "en alquiler de carpas/eventos es limitada, ya que la mayoría de los "
                    "sistemas de inventario están orientados a retail o manufactura. "
                    "SOLUCIÓN:                       "
                    "Se amplió la investigación a sistemas ERP genéricos y se "
                    "extrajeron patrones aplicables al negocio de alquiler. "
                    "APRENDIZAJE:"
                    "El flujo operativo del negocio tiene variaciones según "
                    "el tipo de evento, lo que añade complejidad al modelado."
                ),
            }
        ],
    },
    {
        "numero": 2,
        "periodo": "18/10/2025 - 31/10/2025",
        "fecha_entrega": datetime(2025, 10, 31),
        "actividades": [
            {
                "descripcion": (
                    "SEMANA 3-4: LEVANTAMIENTO DE REQUISITOS Y CASOS DE USO"
                    "                    "
                    "Actividades realizadas:"
                    "                                                                                      "
                    "1. Levantamiento de requisitos funcionales: Se definieron los módulos "
                    "principales del sistema:   "
                    "- Módulo de Inventario: gestión de elementos individuales, series y lotes   "
                    "- Módulo de Productos: categorías jerárquicas y elementos compuestos (plantillas)   "
                    "- Módulo de Alquileres: cotizaciones, eventos y control de disponibilidad   "
                    "- Módulo de Operaciones: flujos de montaje, desmontaje y checklist   "
                    "- Módulo de Configuración: ubicaciones, ciudades y tarifas de transporte"
                    "                                                                                "
                    "2. Definición de requisitos no funcionales: Se establecieron requisitos de "
                    "rendimiento (respuesta < 2s), usabilidad (diseño responsive), seguridad "
                    "y escalabilidad."
                    "                                                                                                                   "
                    "3. Elaboración de casos de uso: Se diseñaron los diagramas de casos de uso "
                    "para cada módulo, identificando actores principales (administrador, operario) "
                    "y flujos alternativos."
                    "                                                "
                    "4. Selección de metodología: Se decidió utilizar una metodología ágil basada "
                    "en iteraciones cortas, priorizando la entrega incremental de funcionalidades."
                ),
                "fecha_inicio": datetime(2025, 10, 18),
                "fecha_fin": datetime(2025, 10, 31),
                "evidencia": (
                    "Documento de especificación de requisitos con 25 requisitos "
                    "funcionales y 8 no funcionales"
                    "                                                      "
                    "- 5 diagramas de casos de uso (uno por módulo)"
                    "                                                    "
                    "- Priorización de módulos: primero Inventario, luego Productos, "
                    "Alquileres, Operaciones y Configuración"
                ),
                "observaciones": (
                    "DIFICULTAD PRINCIPAL:     "
                    "Definir el nivel de granularidad del seguimiento "
                    "de inventario (por unidad individual vs. por lote) fue complejo, ya que "
                    "algunos elementos como sillas se manejan por lote y otros como carpas "
                    "necesitan seguimiento individual por serie. "
                    "SOLUCIÓN:                       "
                    "Investigué sobre sistemas ERP y de inventario. Decidí crear un "
                    "campo \"tipo_gestion\" que puede ser \"serie\" o \"lote\", permitiendo "
                    "gestionar ambos tipos en un mismo sistema. "
                    "APRENDIZAJE:"
                    "La relación entre elementos compuestos (una carpa \"completa\" "
                    "incluye estructura + lona + amarres) requirió un modelado más detallado."
                ),
            }
        ],
    },
    {
        "numero": 3,
        "periodo": "01/11/2025 - 14/11/2025",
        "fecha_entrega": datetime(2025, 11, 14),
        "actividades": [
            {
                "descripcion": (
                    "SEMANA 5-6: DISEÑO DE ARQUITECTURA Y BASE DE DATOS"
                    "                    "
                    "Actividades realizadas:"
                    "                                                                                      "
                    "1. Diseño de la arquitectura: Se definió una arquitectura cliente-servidor con:   "
                    "- Frontend: Aplicación de página única (SPA) con React.js   "
                    "- Backend: API REST con Node.js y Express   "
                    "- Base de datos: MySQL para almacenamiento relacional   "
                    "- Comunicación: API RESTful con JSON"
                    "                                                                                "
                    "2. Selección del stack tecnológico:   "
                    "- Frontend: React 18, Vite, TailwindCSS, React Query, React Router   "
                    "- Backend: Node.js, Express, MySQL2 (driver nativo)   "
                    "- Herramientas: Git/GitHub para control de versiones, ESLint para calidad"
                    "                                                                                                                   "
                    "3. Diseño del modelo de base de datos: Se diseñaron las tablas principales:   "
                    "- categorias: Categorías jerárquicas de productos con soporte de iconos   "
                    "- elementos: Elementos individuales del inventario con cantidad y estado   "
                    "- series: Seguimiento individual de unidades con número de serie   "
                    "- lotes: Agrupación de elementos por cantidad y ubicación   "
                    "- ubicaciones: Ubicaciones físicas (bodega principal, eventos, etc.)   "
                    "- elementos_compuestos: Plantillas de productos compuestos"
                    "                                                "
                    "4. Evaluación de alternativas: Se comparó React vs. Vue.js y MySQL vs. PostgreSQL, "
                    "justificando la elección con base en la curva de aprendizaje, comunidad y "
                    "requisitos del proyecto."
                ),
                "fecha_inicio": datetime(2025, 11, 1),
                "fecha_fin": datetime(2025, 11, 14),
                "evidencia": (
                    "Diagrama de arquitectura del sistema (cliente-servidor con API REST)"
                    "                                                      "
                    "- Diagrama entidad-relación con 12 tablas principales"
                    "                                                    "
                    "- Documento de justificación tecnológica"
                    "                                                     "
                    "- Estructura inicial del repositorio en GitHub"
                ),
                "observaciones": (
                    "DIFICULTAD 1:   "
                    "La decisión entre usar un ORM (Sequelize/Prisma) o "
                    "consultas SQL directas requirió análisis. Se optó por SQL directo con "
                    "MySQL2 para tener mayor control sobre las consultas y evitar la "
                    "abstracción innecesaria de un ORM. "
                    "DIFICULTAD 2:   "
                    "El modelado de elementos compuestos con alternativas "
                    "intercambiables (ej: una carpa puede usar lona tipo A o tipo B) "
                    "requirió un diseño flexible con tablas de relación. "
                    "APRENDIZAJE:"
                    "Un buen análisis de alternativas tecnológicas antes de "
                    "comenzar a codificar evita rehacer trabajo después."
                ),
            }
        ],
    },
    {
        "numero": 4,
        "periodo": "15/11/2025 - 28/11/2025",
        "fecha_entrega": datetime(2025, 11, 28),
        "actividades": [
            {
                "descripcion": (
                    "SEMANA 7-8: CONFIGURACIÓN DEL ENTORNO Y MIGRACIONES"
                    "                    "
                    "Actividades realizadas:"
                    "                                                                                      "
                    "1. Configuración del entorno de desarrollo:   "
                    "- Inicialización del proyecto frontend con Vite + React + TailwindCSS   "
                    "- Configuración del backend con Express y estructura de carpetas por módulos "
                    "(controllers, models, routes)   "
                    "- Configuración de ESLint y reglas de estilo de código   "
                    "- Configuración de variables de entorno para conexión a base de datos"
                    "                                                                                "
                    "2. Creación del sistema de migraciones: Se implementó un sistema de migraciones "
                    "SQL secuenciales para gestionar la evolución del esquema de base de datos de "
                    "forma controlada y reproducible."
                    "                                                                                                                   "
                    "3. Implementación de migraciones base:   "
                    "- Migración de tablas de categorías con soporte de jerarquía (categoría padre)   "
                    "- Migración de elementos, series y lotes   "
                    "- Migración de ubicaciones con tipos predefinidos (bodega, finca, salón, etc.)   "
                    "- Migración de elementos compuestos y sus componentes"
                    "                                                "
                    "4. Estructura del frontend: Se creó la estructura base de la aplicación React con:   "
                    "- Sistema de enrutamiento con React Router   "
                    "- Layout principal con navegación lateral   "
                    "- Configuración de React Query para manejo de estado del servidor"
                ),
                "fecha_inicio": datetime(2025, 11, 15),
                "fecha_fin": datetime(2025, 11, 28),
                "evidencia": (
                    "Proyecto frontend funcional con navegación básica y layout responsivo"
                    "                                                      "
                    "- Backend con API base y conexión a MySQL configurada"
                    "                                                    "
                    "- 10 archivos de migración SQL creados y ejecutados correctamente"
                    "                                                     "
                    "- Repositorio Git configurado con ramas de desarrollo"
                ),
                "observaciones": (
                    "DIFICULTAD 1:   "
                    "La configuración de CORS entre frontend (puerto 5173) y "
                    "backend (puerto 3000) requirió ajustes para permitir las peticiones "
                    "cross-origin durante el desarrollo. "
                    "DIFICULTAD 2:   "
                    "Las migraciones debieron ser diseñadas como idempotentes "
                    "(usando IF NOT EXISTS) para poder re-ejecutarse sin errores en "
                    "diferentes entornos. "
                    "APRENDIZAJE:"
                    "Aprendí la importancia de diseñar migraciones que sean "
                    "seguras de ejecutar múltiples veces."
                ),
            }
        ],
    },
    {
        "numero": 5,
        "periodo": "29/11/2025 - 12/12/2025",
        "fecha_entrega": datetime(2025, 12, 12),
        "actividades": [
            {
                "descripcion": (
                    "SEMANA 9-10: MÓDULO DE INVENTARIO COMPLETO"
                    "                    "
                    "Actividades realizadas:"
                    "                                                                                      "
                    "1. Módulo de Inventario - Backend:   "
                    "- Implementación de controladores CRUD para elementos, series y lotes   "
                    "- Creación de modelos con consultas SQL para cada entidad   "
                    "- Endpoints para obtener elementos con sus series/lotes asociados   "
                    "- Validación de datos de entrada en las rutas"
                    "                                                                                "
                    "2. Módulo de Inventario - Frontend:   "
                    "- Página principal de inventario con listado de elementos en tarjetas   "
                    "- Formularios modales para crear y editar elementos   "
                    "- Vista de detalle de elemento con tabs para series y lotes   "
                    "- Formularios para agregar series (número de serie único) y lotes (cantidad y ubicación)"
                    "                                                                                                                   "
                    "3. Sistema de categorías:   "
                    "- CRUD completo de categorías de productos   "
                    "- Selector de iconos (emojis y Lucide icons) para identificación visual   "
                    "- Tarjetas de categoría con opciones de editar y eliminar"
                    "                                                "
                    "4. Gestión de ubicaciones:   "
                    "- Módulo de ubicaciones accesible desde el dashboard   "
                    "- Funcionalidad de ubicación principal (bodega base)   "
                    "- Tipos de ubicación predefinidos: bodega, finca, salón, hotel, hacienda, club, etc."
                ),
                "fecha_inicio": datetime(2025, 11, 29),
                "fecha_fin": datetime(2025, 12, 12),
                "evidencia": (
                    "Módulo de inventario completamente funcional con CRUD de elementos, "
                    "series y lotes"
                    "                                                      "
                    "- 4 vistas principales implementadas: dashboard, lista de elementos, "
                    "detalle de elemento, gestión de ubicaciones"
                    "                                                    "
                    "- Sistema de categorías con soporte visual de iconos"
                    "                                                     "
                    "- Buscador global integrado en el módulo de inventario"
                ),
                "observaciones": (
                    "DIFICULTAD 1:   "
                    "El renderizado de tarjetas con valor 0 (cero) en React "
                    "mostraba el número en pantalla. Se debió a la evaluación de expresiones "
                    "condicionales con && que tratan el 0 como falsy pero lo renderizan. "
                    "DIFICULTAD 2:   "
                    "La invalidación de queries de React Query al eliminar "
                    "categorías no funcionaba correctamente, requiriendo ajustes en las "
                    "keys de caché. "
                    "APRENDIZAJE:"
                    "En React, usar {count > 0 && ...} en lugar de "
                    "{count && ...} para evitar renderizar el número 0."
                ),
            }
        ],
    },
    {
        "numero": 6,
        "periodo": "13/12/2025 - 26/12/2025",
        "fecha_entrega": datetime(2025, 12, 26),
        "actividades": [
            {
                "descripcion": (
                    "SEMANA 11-12: ELEMENTOS COMPUESTOS Y REORGANIZACIÓN BACKEND"
                    "                    "
                    "Actividades realizadas:"
                    "                                                                                      "
                    "1. Elementos compuestos - Backend:   "
                    "- Implementación del modelo de datos para elementos compuestos: una plantilla "
                    "puede contener múltiples componentes del inventario   "
                    "- Soporte para grupos de componentes alternativos: una carpa puede "
                    "usar lona tipo A o tipo B como alternativa intercambiable   "
                    "- Endpoints CRUD para crear, listar, editar y eliminar plantillas compuestas   "
                    "- Consultas optimizadas para obtener plantillas con sus componentes y conteos"
                    "                                                                                "
                    "2. Elementos compuestos - Frontend:   "
                    "- Formulario multi-paso: paso 1 (datos generales), paso 2 (selección de componentes), "
                    "paso 3 (resumen)   "
                    "- Vista de listado con tarjetas que muestran el conteo de componentes por plantilla   "
                    "- Modal de edición con las mismas funcionalidades del formulario de creación"
                    "                                                                                                                   "
                    "3. Reorganización del backend:   "
                    "- Reestructuración completa en módulos separados: Inventario, Productos y Alquileres   "
                    "- Cada módulo contiene sus propios controllers, models y routes   "
                    "- Simplificación del archivo server.js de ~40 importaciones a una por módulo"
                    "                                                "
                    "4. Mejoras al módulo de inventario:   "
                    "- Corrección del renderizado de 0 en tarjetas de ubicación   "
                    "- Nuevos tipos de ubicación para eventos (hotel, hacienda, club campestre)   "
                    "- Corrección de la integración del selector de ubicaciones"
                ),
                "fecha_inicio": datetime(2025, 12, 13),
                "fecha_fin": datetime(2025, 12, 26),
                "evidencia": (
                    "Módulo de elementos compuestos funcional con soporte de alternativas"
                    "                                                      "
                    "- Backend reorganizado en 3 módulos principales con arquitectura limpia"
                    "                                                    "
                    "- 12 PRs mergeados durante este periodo (#30 a #42)"
                    "                                                     "
                    "- Correcciones de bugs críticos en el módulo de inventario"
                ),
                "observaciones": (
                    "DIFICULTAD 1:   "
                    "La simplificación del modelo de datos para elementos "
                    "compuestos requirió múltiples iteraciones. El diseño original era "
                    "demasiado complejo y se simplificó para mantener la usabilidad. "
                    "DIFICULTAD 2:   "
                    "La reorganización del backend en módulos causó conflictos "
                    "en las rutas de importación que debieron resolverse cuidadosamente. "
                    "APRENDIZAJE:"
                    "Reorganizar la arquitectura temprano facilita el "
                    "mantenimiento a largo plazo."
                ),
            }
        ],
    },
    {
        "numero": 7,
        "periodo": "27/12/2025 - 09/01/2026",
        "fecha_entrega": datetime(2026, 1, 9),
        "actividades": [
            {
                "descripcion": (
                    "SEMANA 13-14: COTIZACIONES, CALENDARIO Y DISPONIBILIDAD"
                    "                    "
                    "Actividades realizadas:"
                    "                                                                                      "
                    "1. Módulo de cotizaciones:   "
                    "- Implementación completa del flujo: creación, edición, vista previa y generación PDF   "
                    "- Vista previa estilo PDF integrada en la interfaz web   "
                    "- Menú kebab (3 puntos) para acciones rápidas en cada cotización   "
                    "- Separación de hooks de cotizaciones en archivos individuales"
                    "                                                                                "
                    "2. Tarifas de transporte:   "
                    "- Tabla maestra de ciudades como entidad independiente   "
                    "- Integración de tarifas de transporte en el formulario de ciudades   "
                    "- Migración completa a ciudad_id como clave foránea   "
                    "- Eliminación del módulo redundante de tarifas que estaba duplicado"
                    "                                                                                                                   "
                    "3. Sistema de calendario:   "
                    "- Integración de FullCalendar con vistas de día, semana, mes y lista   "
                    "- Arquitectura modular del calendario para reutilización   "
                    "- Modal resumen con productos al hacer clic en un evento del calendario"
                    "                                                "
                    "4. Disponibilidad en tiempo real:   "
                    "- Sistema de verificación de disponibilidad de productos por rango de fechas   "
                    "- Debounce en llamadas API para evitar múltiples peticiones simultáneas   "
                    "- Uso de cantidad de tabla elementos como fallback cuando no hay series"
                    "                                                                                "
                    "5. Mejoras generales:   "
                    "- Navegación jerárquica por niveles en productos   "
                    "- Categorías jerárquicas con padre/hijo   "
                    "- Selector de productos con búsqueda y filtrado por categorías"
                ),
                "fecha_inicio": datetime(2025, 12, 27),
                "fecha_fin": datetime(2026, 1, 9),
                "evidencia": (
                    "Módulo de cotizaciones completo con generación de PDF"
                    "                                                      "
                    "- Sistema de calendario integrado con FullCalendar"
                    "                                                    "
                    "- Verificación de disponibilidad funcionando en tiempo real"
                    "                                                     "
                    "- 12 PRs mergeados (#47 a #59)"
                    "                                          "
                    "- Refactorización de la navegación con categorías jerárquicas"
                ),
                "observaciones": (
                    "DIFICULTAD 1:   "
                    "Las múltiples llamadas a la API de disponibilidad "
                    "causaban problemas de rendimiento. Se implementó debounce y se "
                    "deshabilitaron retries en mutaciones. "
                    "DIFICULTAD 2:   "
                    "La migración a ciudad_id requirió un script de migración "
                    "cuidadoso para no perder datos existentes. "
                    "APRENDIZAJE:"
                    "El debounce es esencial en llamadas API frecuentes "
                    "para evitar sobrecarga del servidor."
                ),
            }
        ],
    },
    {
        "numero": 8,
        "periodo": "10/01/2026 - 23/01/2026",
        "fecha_entrega": datetime(2026, 1, 23),
        "actividades": [
            {
                "descripcion": (
                    "SEMANA 15-16: MÓDULO DE OPERACIONES E IMÁGENES"
                    "                    "
                    "Actividades realizadas:"
                    "                                                                                      "
                    "1. Módulo de operaciones:   "
                    "- Implementación del flujo completo: cargue → montaje → desmontaje → retorno → descargue   "
                    "- Órdenes de trabajo con asignación de empleados y vehículos   "
                    "- Historial de estados por orden de trabajo   "
                    "- Checklists de cargue y descargue para verificación de equipos"
                    "                                                                                "
                    "2. Soporte de imágenes:   "
                    "- Upload de imágenes para elementos y productos de alquiler usando Multer   "
                    "- Visualización de fotos en modal durante el proceso de cargue para "
                    "verificación visual   "
                    "- Almacenamiento de imágenes en el servidor con rutas relativas"
                    "                                                                                                                   "
                    "3. Mejoras en cotizaciones:   "
                    "- Fechas de montaje y desmontaje como campos independientes en la cotización   "
                    "- Contexto de alquiler visible en series y lotes (saber qué está alquilado "
                    "y para qué evento)   "
                    "- Corrección de módulos de edición de series y lotes"
                    "                                                "
                    "4. Auto-generación de lotes:   "
                    "- Número de lote generado automáticamente basado en la fecha de creación   "
                    "- Rediseño de la página de detalle de elemento con stat cards como filtros interactivos   "
                    "- Unificación del estado 'nuevo' dentro de 'bueno' para simplificar el flujo"
                    "                                                                                "
                    "5. Limpieza y mantenimiento:   "
                    "- Script SQL para limpiar datos de prueba de la base de datos   "
                    "- Corrección de subtotal de cotización que no incluía recargos   "
                    "- Migración 27 hecha idempotente para compatibilidad"
                ),
                "fecha_inicio": datetime(2026, 1, 10),
                "fecha_fin": datetime(2026, 1, 23),
                "evidencia": (
                    "Módulo de operaciones funcional con flujo completo de 5 etapas"
                    "                                                      "
                    "- Soporte de imágenes implementado para elementos y productos"
                    "                                                    "
                    "- Auto-generación de números de lote"
                    "                                                     "
                    "- Rediseño de la vista de detalle de elemento"
                    "                                          "
                    "- 8 PRs mergeados (#60 a #67)"
                ),
                "observaciones": (
                    "DIFICULTAD 1:   "
                    "El envío del campo cantidad al crear elementos no se "
                    "estaba realizando correctamente, causando que los elementos se "
                    "crearan sin stock inicial. "
                    "DIFICULTAD 2:   "
                    "Los iconos SVG no se renderizaban correctamente en "
                    "algunos componentes; se debió crear un componente IconoCategoria "
                    "reutilizable. "
                    "APRENDIZAJE:"
                    "Crear componentes reutilizables para lógica repetida "
                    "reduce errores y facilita el mantenimiento."
                ),
            }
        ],
    },
    {
        "numero": 9,
        "periodo": "24/01/2026 - 06/02/2026",
        "fecha_entrega": datetime(2026, 2, 6),
        "actividades": [
            {
                "descripcion": (
                    "SEMANA 17-18: FUNCIONALIDADES AVANZADAS"
                    "                    "
                    "Actividades realizadas:"
                    "                                                                                      "
                    "1. Cronómetros en tiempo real:   "
                    "- Implementación de temporizadores en vivo para operaciones de montaje y desmontaje   "
                    "- Los cronómetros inician cuando se cambia el estado de la orden y se detienen al finalizar   "
                    "- Modales de confirmación estilizados para cambios de estado   "
                    "- Auto-creación de la tabla orden_trabajo_historial_estados"
                    "                                                                                "
                    "2. Gestión de depósitos:   "
                    "- Campo de valor de depósito agregado a las cotizaciones   "
                    "- Toggle de cobro de depósito (cobrar o no cobrar)   "
                    "- Uso de valores del backend (resumen) en vez de recalcular subtotales   "
                    "- Botón \"Ver PDF\" integrado en la cotización"
                    "                                                                                                                   "
                    "3. Historial de eventos por cliente:   "
                    "- Nueva página de eventos completados con historial de productos alquilados   "
                    "- Filtro para mostrar solo eventos activos en la lista de cotizaciones   "
                    "- Visualización de productos alquilados y cantidades en el historial"
                    "                                                "
                    "4. Funcionalidad de repetir evento:   "
                    "- Botón \"Repetir evento\" que crea nueva cotización con los mismos productos   "
                    "- Modal de formulario con campos de fecha resaltados   "
                    "- Conversión de valores DECIMAL de MySQL a números en JavaScript"
                    "                                                                                "
                    "5. Mejoras en operaciones:   "
                    "- Etapas en_retorno y descargue agregadas al flujo de desmontaje   "
                    "- Renombre de \"Checklist de Descargue\" a \"Recogida\" y nuevo \"Checklist en Bodega\"   "
                    "- Mejora de temporizadores para montaje y desmontaje"
                ),
                "fecha_inicio": datetime(2026, 1, 24),
                "fecha_fin": datetime(2026, 2, 6),
                "evidencia": (
                    "Cronómetros en tiempo real funcionando para montaje y desmontaje"
                    "                                                      "
                    "- Sistema de depósitos integrado en cotizaciones"
                    "                                                    "
                    "- Historial completo de eventos por cliente con opción de repetir"
                    "                                                     "
                    "- Flujo de desmontaje ampliado con 2 nuevas etapas"
                    "                                          "
                    "- 6 PRs mergeados (#77 a #82)"
                ),
                "observaciones": (
                    "DIFICULTAD 1:   "
                    "Los hooks de contexto nunca se ejecutaban debido a una "
                    "comparación estricta (===) con booleanos de MySQL, que devuelve 0 y 1 "
                    "en lugar de true y false. Se solucionó convirtiendo a booleano. "
                    "DIFICULTAD 2:   "
                    "El valor requiere_series sin convertir a booleano provocaba "
                    "que se renderizara un 0 en pantalla. "
                    "APRENDIZAJE:"
                    "MySQL devuelve 0/1 para booleanos, no true/false. "
                    "Siempre convertir explícitamente en JavaScript."
                ),
            }
        ],
    },
    {
        "numero": 10,
        "periodo": "07/02/2026 - 20/02/2026",
        "fecha_entrega": datetime(2026, 2, 20),
        "actividades": [
            {
                "descripcion": (
                    "SEMANA 19-20: CORRECCIÓN DE BUGS Y REORGANIZACIÓN FINAL"
                    "                    "
                    "Actividades realizadas:"
                    "                                                                                      "
                    "1. Corrección del bug de stock cero:   "
                    "- Se identificó que las cotizaciones descontaban el stock al crearlas y "
                    "nuevamente al confirmar el alquiler, provocando un doble descuento   "
                    "- Se corrigió el mapeo de disponibilidad por producto en el selector   "
                    "- Pruebas exhaustivas del flujo completo: crear cotización → confirmar "
                    "alquiler → verificar stock → devolver → verificar restauración de stock"
                    "                                                                                "
                    "2. Rediseño de la interfaz de inventario:   "
                    "- Nueva interfaz del módulo con diseño más limpio y moderno   "
                    "- Tarjetas de productos con mejor visualización de stock y estados   "
                    "- Mejoras en la experiencia de usuario para navegación entre módulos"
                    "                                                                                                                   "
                    "3. Reorganización completa del proyecto:   "
                    "- Backend: Creación de 7 módulos independientes (auth, inventario, productos, "
                    "alquileres, clientes, operaciones, configuracion)   "
                    "- Frontend: Reorganización de 247 archivos en 9 módulos + shared   "
                    "- Simplificación del servidor: server.js de ~40 líneas de importación a una por módulo   "
                    "- Actualización de todas las rutas de importación"
                    "                                                "
                    "4. Resolución de conflictos post-reorganización:   "
                    "- Corrección de rutas de importación de ClienteModel y ConfiguracionModel   "
                    "- Merge de conflictos entre la rama de reorganización y main"
                ),
                "fecha_inicio": datetime(2026, 2, 7),
                "fecha_fin": datetime(2026, 2, 20),
                "evidencia": (
                    "Bug crítico de doble descuento de stock corregido"
                    "                                                      "
                    "- Interfaz de inventario rediseñada con mejor UX"
                    "                                                    "
                    "- Proyecto completamente reorganizado en arquitectura modular:"
                    "  7 módulos backend, 9 módulos frontend + shared, 261 endpoints API"
                    "                                                     "
                    "- 3 PRs mergeados (#83, #84, #85)"
                    "                                          "
                    "ESTADO FINAL: 319 archivos JS/JSX, 106 componentes React, "
                    "36 hooks, 31 tablas BD, 29 migraciones, 383+ commits, 85 PRs"
                ),
                "observaciones": (
                    "DIFICULTAD PRINCIPAL:     "
                    "La reorganización de 247 archivos fue la tarea más "
                    "compleja del proyecto. Cada archivo movido requería actualizar todas sus "
                    "importaciones y las importaciones de los archivos que lo referenciaban. "
                    "DIFICULTAD 2:   "
                    "Algunos controladores seguían importando modelos con rutas "
                    "relativas antiguas (../../models/) en lugar de las nuevas rutas modulares. "
                    "APRENDIZAJE:"
                    "Una buena organización modular desde el inicio del proyecto "
                    "habría evitado esta reorganización masiva. Lección para futuros proyectos."
                ),
            }
        ],
    },
]


# ─── Estilos exactos del formato original ─────────────────────────────────────

THIN = Side(style="thin")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Fuentes - El original usa Calibri 12 para todo, Calibri 9 para modalidades
FONT_12 = Font(name="Calibri", size=12)
FONT_12_BOLD = Font(name="Calibri", size=12, bold=True)
FONT_11 = Font(name="Calibri", size=11)
FONT_9_BOLD = Font(name="Calibri", size=9, bold=True)
FONT_X = Font(name="Calibri", size=36, bold=True)  # La "x" de la modalidad

# Alineaciones
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_CENTER_WRAP = Alignment(horizontal="center", wrap_text=True)
ALIGN_LEFT_WRAP = Alignment(horizontal="left", wrap_text=True)
ALIGN_LEFT_CENTER = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_V_CENTER_WRAP = Alignment(vertical="center", wrap_text=True)
ALIGN_V_TOP_WRAP = Alignment(vertical="top", wrap_text=True)

# Anchos de columna del original
COL_WIDTHS = {
    "A": 2.29, "B": 5.71, "C": 8.14, "D": 5.71, "E": 7.14,
    "F": 7.71, "G": 5.71, "H": 8.57, "I": 7.14, "J": 5.71,
    "K": 5.71, "L": 5.71, "M": 5.71, "N": 9.00, "O": 5.71,
    "P": 5.71, "Q": 5.86, "R": 5.71, "S": 7.14, "T": 5.71,
    "U": 5.71, "V": 5.71, "W": 5.71, "X": 5.29, "Y": 5.71,
}

# Alturas de fila del original (filas fijas del encabezado)
ROW_HEIGHTS = {
    1: 35.25, 2: 35.25, 3: 18.00, 4: 18.75,
    5: 32.25, 6: 32.25, 7: 32.25, 8: 16.50,
    9: 18.00, 10: 18.00, 11: 9.95,
    12: 18.00, 13: 18.00, 14: 9.95,
    15: 18.00, 16: 31.50, 17: 45.75, 18: 9.95,
    19: 18.00, 20: 18.00, 21: 9.75,
    22: 18.00, 23: 18.00, 24: 9.75,
    25: 48.00, 26: 18.00,
}


def apply_border_range(ws, min_row, max_row, min_col, max_col):
    """Apply thin border to a range of cells."""
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER


def merge_write(ws, min_row, min_col, max_row, max_col, value,
                font=FONT_12, alignment=ALIGN_CENTER_WRAP, number_format=None):
    """Merge cells and write value with formatting, then apply borders."""
    if min_row != max_row or min_col != max_col:
        ws.merge_cells(
            start_row=min_row, start_column=min_col,
            end_row=max_row, end_column=max_col
        )
    cell = ws.cell(row=min_row, column=min_col, value=value)
    cell.font = font
    cell.alignment = alignment
    if number_format:
        cell.number_format = number_format
    apply_border_range(ws, min_row, max_row, min_col, max_col)


def create_bitacora_sheet(wb, data):
    """Create a single bitácora sheet replicating the exact SENA format."""
    num = data["numero"]
    ws = wb.create_sheet(title="Bitácora {:02d}".format(num))

    # ── Column widths ──
    for letter, width in COL_WIDTHS.items():
        ws.column_dimensions[letter].width = width

    # ── Row heights (fixed header rows) ──
    for row_num, height in ROW_HEIGHTS.items():
        ws.row_dimensions[row_num].height = height

    # ── Row 1-2: Header area + Version (B1:V2 merged empty, W1:Y1 version) ──
    merge_write(ws, 1, 2, 2, 22, None, FONT_12, ALIGN_CENTER)  # B1:V2 empty
    merge_write(ws, 1, 23, 1, 25, "Versión: 03", FONT_11, ALIGN_CENTER)
    merge_write(ws, 2, 23, 2, 25, "Código: \nGFPI-F-147", FONT_11, ALIGN_CENTER)

    # ── Row 3: Title line 1 ──
    merge_write(ws, 3, 2, 3, 25,
                "Proceso Gestión de Formación Profesional Integral",
                FONT_12_BOLD, ALIGN_CENTER)

    # ── Row 4: Title line 2 ──
    merge_write(ws, 4, 2, 4, 25,
                " Formato Bitácora seguimiento Etapa productiva",
                FONT_12_BOLD, ALIGN_CENTER)

    # ── Row 5: Regional ──
    merge_write(ws, 5, 2, 5, 25, DATOS["regional"],
                FONT_12_BOLD, ALIGN_CENTER_WRAP)

    # ── Row 6: Centro ──
    merge_write(ws, 6, 2, 6, 25, DATOS["centro"],
                FONT_12_BOLD, ALIGN_CENTER_WRAP)

    # ── Row 7: Title bar ──
    merge_write(ws, 7, 2, 7, 25,
                "BITÁCORA DE SEGUIMIENTO ETAPA PRODUCTIVA",
                FONT_12_BOLD, ALIGN_CENTER_WRAP)

    # ── Row 8: Empty (B8:Y8 merged) ──
    merge_write(ws, 8, 2, 8, 25, None, FONT_12, ALIGN_CENTER)

    # ── Row 9: Company labels ──
    merge_write(ws, 9, 2, 9, 12,
                "Nombre de la empresa donde está realizando la etapa productiva",
                FONT_12_BOLD, ALIGN_CENTER)
    merge_write(ws, 9, 13, 9, 15, "NIT", FONT_12_BOLD, ALIGN_CENTER)
    merge_write(ws, 9, 16, 9, 19, "BITACORA  N°", FONT_12_BOLD, ALIGN_CENTER)
    merge_write(ws, 9, 20, 9, 25, "Período ", FONT_12_BOLD, ALIGN_CENTER)

    # ── Row 10: Company values ──
    merge_write(ws, 10, 2, 10, 12, DATOS["empresa"],
                FONT_12, ALIGN_CENTER_WRAP)
    merge_write(ws, 10, 13, 10, 15, DATOS["nit"],
                FONT_12_BOLD, ALIGN_CENTER)
    merge_write(ws, 10, 16, 10, 19, num,
                FONT_12_BOLD, ALIGN_CENTER)
    merge_write(ws, 10, 20, 10, 25, data["periodo"],
                FONT_12, ALIGN_CENTER_WRAP)

    # ── Row 11: Empty spacer ──

    # ── Row 12: Supervisor labels ──
    merge_write(ws, 12, 2, 12, 10,
                "Nombre del jefe inmediato/Responsable",
                FONT_12_BOLD, ALIGN_CENTER)
    merge_write(ws, 12, 11, 12, 16, "Teléfono de contacto",
                FONT_12_BOLD, ALIGN_CENTER)
    merge_write(ws, 12, 17, 12, 25, "Correo electrónico",
                FONT_12_BOLD, ALIGN_CENTER)

    # ── Row 13: Supervisor values ──
    merge_write(ws, 13, 2, 13, 10, DATOS["jefe_nombre"],
                FONT_12, ALIGN_LEFT_WRAP)
    merge_write(ws, 13, 11, 13, 16, DATOS["jefe_telefono"],
                FONT_12, ALIGN_LEFT_WRAP)
    merge_write(ws, 13, 17, 13, 25, DATOS["jefe_correo"],
                FONT_11, ALIGN_LEFT_WRAP)

    # ── Row 14: Empty spacer ──

    # ── Row 15: Modality label ──
    merge_write(ws, 15, 2, 15, 25,
                'Seleccione con una "X" el tipo de modalidad de etapa productiva',
                FONT_12_BOLD, ALIGN_LEFT_WRAP)

    # ── Row 16-17: Modalities (each spans 2 rows) ──
    modalities = [
        (2, 3, "CONTRATO DE APRENDIZAJE"),
        (4, 4, ""),       # separator
        (5, 6, "VÍNCULO LABORAL O CONTRACTUAL"),
        (7, 7, "x"),      # X mark
        (8, 9, "PROYECTO PRODUCTIVO"),
        (10, 10, ""),      # separator
        (11, 13, "APOYO A UNA UNIDAD PRODUCTIVA FAMILIAR"),
        (14, 14, ""),      # separator
        (15, 17, "APOYO A INSTITUCIÓN ESTATAL NACIONAL,TERRITORIAL, O A UNA ONG, O A ENTIDAD SIN ANIMO DE LUCRO"),
        (18, 18, ""),      # separator
        (19, 20, "MONITORIA"),
        (21, 21, ""),      # separator
        (22, 24, "PASANTIA"),
        (25, 25, ""),      # separator
    ]
    for col_start, col_end, label in modalities:
        if label == "x":
            merge_write(ws, 16, col_start, 17, col_end, "x",
                        FONT_X, ALIGN_CENTER)
        else:
            merge_write(ws, 16, col_start, 17, col_end, label if label else None,
                        FONT_9_BOLD, ALIGN_CENTER)

    # ── Row 18: Empty spacer ──

    # ── Row 19: Student labels ──
    merge_write(ws, 19, 2, 19, 9, "Nombre del aprendiz",
                FONT_12_BOLD, ALIGN_CENTER_WRAP)
    merge_write(ws, 19, 10, 19, 14, "Documento Id.",
                FONT_12_BOLD, ALIGN_CENTER_WRAP)
    merge_write(ws, 19, 15, 19, 19, "Teléfono de contacto",
                FONT_12_BOLD, ALIGN_CENTER_WRAP)
    merge_write(ws, 19, 20, 19, 25, "Correo electrónico institucional",
                FONT_12_BOLD, ALIGN_CENTER_WRAP)

    # ── Row 20: Student values ──
    merge_write(ws, 20, 2, 20, 9, DATOS["aprendiz_nombre"],
                FONT_12, ALIGN_LEFT_WRAP)
    merge_write(ws, 20, 10, 20, 14, DATOS["aprendiz_doc"],
                FONT_12, ALIGN_LEFT_WRAP)
    merge_write(ws, 20, 15, 20, 19, DATOS["aprendiz_tel"],
                FONT_12_BOLD, ALIGN_CENTER_WRAP)
    merge_write(ws, 20, 20, 20, 25, DATOS["aprendiz_correo"],
                FONT_11, ALIGN_CENTER_WRAP)

    # ── Row 21: Empty spacer ──

    # ── Row 22: Program labels ──
    merge_write(ws, 22, 2, 22, 10, "Número de ficha",
                FONT_12_BOLD, ALIGN_CENTER_WRAP)
    merge_write(ws, 22, 11, 22, 25, "Programa de formación",
                FONT_12_BOLD, ALIGN_CENTER)

    # ── Row 23: Program values ──
    merge_write(ws, 23, 2, 23, 10, DATOS["ficha"],
                FONT_12, ALIGN_CENTER_WRAP)
    merge_write(ws, 23, 11, 23, 25, DATOS["programa"],
                FONT_12, ALIGN_CENTER_WRAP)

    # ── Row 24: Empty spacer ──

    # ── Row 25: Activities table header ──
    merge_write(ws, 25, 2, 25, 10,
                "DESCRIPCIÓN DE LA ACTIVIDAD"
                "                                               "
                "(Ingrese cuantas filas sean necesarias)",
                FONT_12_BOLD, ALIGN_CENTER)
    merge_write(ws, 25, 11, 25, 12, "FECHA INICIO",
                FONT_12_BOLD, ALIGN_CENTER)
    merge_write(ws, 25, 13, 25, 14, "FECHA \nFIN",
                FONT_12_BOLD, ALIGN_CENTER)
    merge_write(ws, 25, 15, 25, 20, "EVIDENCIA DE CUMPLIMIENTO",
                FONT_12_BOLD, ALIGN_CENTER)
    merge_write(ws, 25, 21, 25, 25,
                "OBSERVACIONES, INASISTENCIAS  Y/O DIFICULTADES PRESENTADAS",
                FONT_12_BOLD, ALIGN_CENTER)

    # ── Activities: each spans 24 rows (like original B26:J49) ──
    current_row = 26
    for i, act in enumerate(data["actividades"]):
        act_rows = 24  # Each activity block = 24 rows (matching original)
        end_row = current_row + act_rows - 1

        # Description
        merge_write(ws, current_row, 2, end_row, 10,
                    act["descripcion"],
                    FONT_12, ALIGN_LEFT_CENTER)

        # Fecha inicio (date format d/mm/yyyy)
        merge_write(ws, current_row, 11, end_row, 12,
                    act["fecha_inicio"],
                    FONT_12_BOLD, ALIGN_CENTER, "d/mm/yyyy;@")

        # Fecha fin (date format d/mm/yyyy)
        merge_write(ws, current_row, 13, end_row, 14,
                    act["fecha_fin"],
                    FONT_12_BOLD, ALIGN_CENTER, "d/mm/yyyy;@")

        # Evidencia
        merge_write(ws, current_row, 15, end_row, 20,
                    act["evidencia"],
                    FONT_12, ALIGN_V_CENTER_WRAP)

        # Observaciones
        merge_write(ws, current_row, 21, end_row, 25,
                    act["observaciones"],
                    FONT_12, ALIGN_V_TOP_WRAP)

        # Set row heights for activity rows
        for r in range(current_row, end_row + 1):
            ws.row_dimensions[r].height = 15.00

        current_row = end_row + 1

    # ── Fill remaining rows up to row 108 with empty bordered cells ──
    # (Original has empty rows between activities and signatures)
    while current_row <= 108:
        ws.row_dimensions[current_row].height = 15.00
        current_row += 1

    # ── Row 109: Reminder ──
    ws.row_dimensions[109].height = 15.00
    merge_write(ws, 109, 2, 109, 25,
                "Aprendiz: recuerde diligenciar completamente el informe y "
                "entregarlo o subirlo  al espacio asignado para este.",
                FONT_12_BOLD, ALIGN_CENTER_WRAP)

    # ── Rows 110-111: Empty + Student name ──
    ws.row_dimensions[110].height = 15.00
    ws.row_dimensions[111].height = 15.00
    merge_write(ws, 111, 3, 111, 9, "ANDERSON MORENO",
                FONT_12_BOLD, ALIGN_CENTER_WRAP)

    # ── Row 112: Date ──
    ws.row_dimensions[112].height = 15.00
    merge_write(ws, 112, 21, 112, 24, data["fecha_entrega"],
                FONT_12, ALIGN_CENTER_WRAP, "dd/mm/yyyy;@")

    # ── Row 113: Labels ──
    ws.row_dimensions[113].height = 15.00
    merge_write(ws, 113, 2, 113, 9, "Nombre del Aprendiz",
                FONT_12_BOLD, ALIGN_CENTER_WRAP)
    merge_write(ws, 113, 11, 113, 17, "Firma del aprendiz",
                FONT_12_BOLD, ALIGN_CENTER_WRAP)
    merge_write(ws, 113, 20, 113, 24, "Fecha entrega bitácora",
                FONT_12_BOLD, ALIGN_CENTER_WRAP)

    # ── Rows 114-116: Empty ──
    for r in range(114, 117):
        ws.row_dimensions[r].height = 15.00

    merge_write(ws, 115, 2, 115, 9, None, FONT_12, ALIGN_CENTER_WRAP)

    # ── Row 117: Instructor + Jefe labels ──
    ws.row_dimensions[117].height = 15.75
    merge_write(ws, 117, 2, 117, 9,
                "Nombre del Instructor de Seguimiento",
                FONT_12_BOLD, ALIGN_CENTER_WRAP)
    merge_write(ws, 117, 11, 117, 17,
                "Firma de instructor de seguimiento",
                FONT_12_BOLD, ALIGN_CENTER_WRAP)
    merge_write(ws, 117, 19, 117, 25,
                "Firma del jefe inmediato (Si es del caso)",
                FONT_12, ALIGN_LEFT_WRAP)

    # ── Row 118: Empty with date label area ──
    ws.row_dimensions[118].height = 23.25

    # ── Row 119-120: Privacy note ──
    ws.row_dimensions[119].height = 18.00
    merge_write(ws, 119, 2, 120, 25,
                "Nota:  LOS DATOS PROPORCIONADOS SERÁN TRATADOS DE ACUERDO CON "
                "LA POLÍTICA DE TRATAMIENTO DE DATOS PERSONALES DEL SENA Y A LA "
                "LEY 1581 DE 2012.",
                FONT_12, ALIGN_LEFT_WRAP)

    # ── Page setup ──
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def create_instrucciones_sheet(wb):
    """Create the 'Instrucciones' sheet (simplified)."""
    ws = wb.create_sheet(title="Instrucciones")

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 40
    for c in range(3, 16):
        ws.column_dimensions[get_column_letter(c)].width = 8

    merge_write(ws, 4, 2, 4, 10,
                "Proceso Gestión de Formación Profesional Integral\n"
                "Formato Bitácora seguimiento Etapa productiva",
                FONT_12_BOLD, ALIGN_CENTER)
    merge_write(ws, 4, 11, 4, 15, "Versión: 03", FONT_11, ALIGN_CENTER)
    merge_write(ws, 5, 11, 5, 15, "Código: \nGFPI-F-147", FONT_11, ALIGN_CENTER)

    merge_write(ws, 6, 2, 6, 15,
                "INSTRUCCIONES PARA EL DILIGENCIAMIENTO DEL FORMATO",
                FONT_12_BOLD, ALIGN_CENTER)
    merge_write(ws, 7, 2, 7, 15, "NO IMPRIMIR", FONT_12_BOLD, ALIGN_CENTER)
    merge_write(ws, 8, 2, 8, 15, "1. Generalidades ", FONT_12_BOLD, ALIGN_LEFT_WRAP)
    ws.row_dimensions[9].height = 80
    merge_write(ws, 9, 2, 9, 15,
                "El formato de Bitácora de Seguimiento de Etapa Productiva se crea como "
                "parte de la Guía Desarrollo Etapa Productiva en el proceso formativo que "
                "se ubica dentro del Proceso Ejecución de la Formación Profesional Integral "
                "(GFPI-P-006), para ser diligenciado por los aprendices y el jefe de la "
                "empresa (si es del caso para las alternativas de etapa productiva de índole "
                "laboral).  Este formato se diligencia según la frecuencia establecida en "
                "el Reglamento del Aprendiz vigente o lineamientos SENA.  Una vez "
                "diligenciado debe entregarlo en físico o subirlo a la plataforma LMS como "
                "evidencia de la ejecución de la etapa productiva por parte de los aprendices "
                "para posterior revisión y retroalimentación por parte de los instructores "
                "de seguimiento asignados por el Centro de Formación.",
                FONT_12, ALIGN_LEFT_WRAP)


def main():
    wb = Workbook()

    create_instrucciones_sheet(wb)

    for data in BITACORAS_DATA:
        create_bitacora_sheet(wb, data)

    # Remove default sheet
    del wb["Sheet"]

    output_path = os.path.join(BITACORAS_DIR, "Bitacoras_SENA_GFPI-F-147.xlsx")
    wb.save(output_path)
    print("Excel generado: {}".format(output_path))
    print("  {} bitacoras en formato SENA GFPI-F-147".format(len(BITACORAS_DATA)))


if __name__ == "__main__":
    main()
