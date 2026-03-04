from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()

bitacoras = [
    {
        "numero": 1,
        "fecha": "4 de octubre de 2025",
        "periodo": "4 de octubre – 17 de octubre de 2025",
        "horas": 20,
        "objetivo": "Realizar la investigación inicial sobre sistemas de gestión de inventario existentes y definir el alcance del proyecto.",
        "actividades": [
            "Investigación de antecedentes: Se revisaron sistemas de gestión de inventario existentes en el mercado, incluyendo soluciones SaaS como Sortly, inFlow Inventory y Zoho Inventory, identificando fortalezas y limitaciones para el contexto de alquiler de carpas y eventos.",
            "Análisis del dominio del negocio: Se realizaron reuniones con el negocio de alquiler de carpas para comprender el flujo operativo completo: recepción de pedidos, cotización, despacho, montaje, desmontaje y retorno al inventario.",
            "Identificación del problema: Se documentó que el negocio actualmente maneja el inventario de manera manual (hojas de cálculo y registros en papel), lo que genera errores en la disponibilidad de productos, pérdida de elementos y dificultad para cotizar eventos.",
            "Revisión de literatura: Se consultaron artículos académicos sobre sistemas de gestión de inventario, metodologías ágiles de desarrollo de software y arquitecturas de aplicaciones web modernas.",
        ],
        "resultados": [
            "Documento de antecedentes con 5 sistemas analizados y sus características principales.",
            "Mapa del flujo operativo del negocio de alquiler de carpas.",
            "Identificación clara de los problemas que el sistema debe resolver: control de stock, trazabilidad por series/lotes, cotizaciones y seguimiento de operaciones.",
        ],
        "dificultades": [
            "La información sobre sistemas especializados en alquiler de carpas/eventos es limitada, ya que la mayoría de los sistemas de inventario están orientados a retail o manufactura.",
            "El flujo operativo del negocio tiene variaciones según el tipo de evento, lo que añade complejidad al modelado.",
        ],
        "plan_siguiente": [
            "Definir los requisitos funcionales y no funcionales del sistema.",
            "Seleccionar la metodología de desarrollo.",
            "Elaborar los casos de uso principales.",
        ],
    },
    {
        "numero": 2,
        "fecha": "18 de octubre de 2025",
        "periodo": "18 de octubre – 31 de octubre de 2025",
        "horas": 22,
        "objetivo": "Definir los requisitos funcionales y no funcionales del sistema y elaborar los diagramas de casos de uso principales.",
        "actividades": [
            "Levantamiento de requisitos funcionales: Se definieron los módulos principales del sistema: Módulo de Inventario (gestión de elementos individuales, series y lotes), Módulo de Productos (categorías jerárquicas y elementos compuestos), Módulo de Alquileres (cotizaciones, eventos y control de disponibilidad), Módulo de Operaciones (flujos de montaje, desmontaje y checklist), Módulo de Configuración (ubicaciones, ciudades y tarifas de transporte).",
            "Definición de requisitos no funcionales: Se establecieron requisitos de rendimiento (respuesta < 2s), usabilidad (diseño responsive), seguridad y escalabilidad.",
            "Elaboración de casos de uso: Se diseñaron los diagramas de casos de uso para cada módulo, identificando actores principales (administrador, operario) y flujos alternativos.",
            "Selección de metodología: Se decidió utilizar una metodología ágil basada en iteraciones cortas, priorizando la entrega incremental de funcionalidades.",
        ],
        "resultados": [
            "Documento de especificación de requisitos con 25 requisitos funcionales y 8 no funcionales.",
            "5 diagramas de casos de uso (uno por módulo).",
            "Priorización de módulos: primero Inventario, luego Productos, Alquileres, Operaciones y Configuración.",
        ],
        "dificultades": [
            "Definir el nivel de granularidad del seguimiento de inventario (por unidad individual vs. por lote) fue complejo, ya que algunos elementos como sillas se manejan por lote y otros como carpas necesitan seguimiento individual por serie.",
            "La relación entre elementos compuestos (una carpa 'completa' incluye estructura + lona + amarres) requirió un modelado más detallado.",
        ],
        "plan_siguiente": [
            "Diseñar la arquitectura técnica del sistema.",
            "Seleccionar el stack tecnológico.",
            "Diseñar el modelo de base de datos.",
        ],
    },
    {
        "numero": 3,
        "fecha": "1 de noviembre de 2025",
        "periodo": "1 de noviembre – 14 de noviembre de 2025",
        "horas": 25,
        "objetivo": "Diseñar la arquitectura del sistema, seleccionar el stack tecnológico y crear el modelo de base de datos.",
        "actividades": [
            "Diseño de la arquitectura: Se definió una arquitectura cliente-servidor con: Frontend (SPA con React.js), Backend (API REST con Node.js y Express), Base de datos (MySQL para almacenamiento relacional), Comunicación (API RESTful con JSON).",
            "Selección del stack tecnológico: Frontend (React 18, Vite, TailwindCSS, React Query/TanStack Query, React Router), Backend (Node.js, Express, MySQL2 driver nativo), Herramientas (Git/GitHub para control de versiones, ESLint para calidad de código).",
            "Diseño del modelo de base de datos: Se diseñaron las tablas principales: categorias (jerárquicas con soporte de iconos), elementos (del inventario con cantidad y estado), series (seguimiento individual con número de serie y ubicación), lotes (agrupación por cantidad y ubicación), ubicaciones (físicas: bodega principal, eventos, etc.), elementos_compuestos (plantillas de productos con componentes).",
            "Evaluación de alternativas: Se comparó React vs. Vue.js y MySQL vs. PostgreSQL, justificando la elección con base en la curva de aprendizaje, comunidad y requisitos del proyecto.",
        ],
        "resultados": [
            "Diagrama de arquitectura del sistema (cliente-servidor con API REST).",
            "Diagrama entidad-relación con 12 tablas principales.",
            "Documento de justificación tecnológica.",
            "Estructura inicial del repositorio en GitHub.",
        ],
        "dificultades": [
            "La decisión entre usar un ORM (Sequelize/Prisma) o consultas SQL directas requirió análisis. Se optó por SQL directo con MySQL2 para tener mayor control sobre las consultas y evitar la abstracción innecesaria de un ORM.",
            "El modelado de elementos compuestos con alternativas intercambiables (ej: una carpa puede usar lona tipo A o tipo B) requirió un diseño flexible con tablas de relación.",
        ],
        "plan_siguiente": [
            "Configurar el entorno de desarrollo (frontend y backend).",
            "Implementar la estructura base del proyecto.",
            "Crear las migraciones de base de datos.",
        ],
    },
    {
        "numero": 4,
        "fecha": "15 de noviembre de 2025",
        "periodo": "15 de noviembre – 28 de noviembre de 2025",
        "horas": 28,
        "objetivo": "Configurar el entorno de desarrollo completo y crear la estructura base del proyecto con las migraciones de base de datos.",
        "actividades": [
            "Configuración del entorno de desarrollo: Inicialización del proyecto frontend con Vite + React + TailwindCSS. Configuración del backend con Express y estructura de carpetas por módulos (controllers, models, routes). Configuración de ESLint y reglas de estilo de código. Configuración de variables de entorno para conexión a base de datos.",
            "Creación del sistema de migraciones: Se implementó un sistema de migraciones SQL secuenciales para gestionar la evolución del esquema de base de datos de forma controlada y reproducible.",
            "Implementación de migraciones base: Migración de tablas de categorías con soporte de jerarquía (categoría padre). Migración de elementos, series y lotes. Migración de ubicaciones con tipos predefinidos (bodega, finca, salón, etc.). Migración de elementos compuestos y sus componentes.",
            "Estructura del frontend: Se creó la estructura base de la aplicación React con: Sistema de enrutamiento con React Router, Layout principal con navegación lateral, Configuración de React Query para manejo de estado del servidor.",
        ],
        "resultados": [
            "Proyecto frontend funcional con navegación básica y layout responsivo.",
            "Backend con API base y conexión a MySQL configurada.",
            "10 archivos de migración SQL creados y ejecutados correctamente.",
            "Repositorio Git configurado con ramas de desarrollo.",
        ],
        "dificultades": [
            "La configuración de CORS entre frontend (puerto 5173) y backend (puerto 3000) requirió ajustes para permitir las peticiones cross-origin durante el desarrollo.",
            "Las migraciones debieron ser diseñadas como idempotentes (usando IF NOT EXISTS) para poder re-ejecutarse sin errores en diferentes entornos.",
        ],
        "plan_siguiente": [
            "Implementar el módulo de inventario: CRUD de elementos, series y lotes.",
            "Crear las interfaces de usuario para gestión de inventario.",
            "Implementar el sistema de categorías de productos.",
        ],
    },
    {
        "numero": 5,
        "fecha": "29 de noviembre de 2025",
        "periodo": "29 de noviembre – 12 de diciembre de 2025",
        "horas": 30,
        "objetivo": "Implementar el módulo de inventario completo con gestión de elementos, series, lotes y categorías.",
        "actividades": [
            "Módulo de Inventario - Backend: Implementación de controladores CRUD para elementos, series y lotes. Creación de modelos con consultas SQL para cada entidad. Endpoints para obtener elementos con sus series/lotes asociados. Validación de datos de entrada en las rutas.",
            "Módulo de Inventario - Frontend: Página principal de inventario con listado de elementos en tarjetas. Formularios modales para crear y editar elementos. Vista de detalle de elemento con tabs para series y lotes. Formularios para agregar series (con número de serie único) y lotes (con cantidad y ubicación).",
            "Sistema de categorías: CRUD completo de categorías de productos. Selector de iconos (emojis y Lucide icons) para identificación visual de categorías. Tarjetas de categoría con opciones de editar y eliminar.",
            "Gestión de ubicaciones: Módulo de ubicaciones accesible desde el dashboard. Funcionalidad de ubicación principal (bodega base). Tipos de ubicación predefinidos: bodega, finca, salón, hotel, hacienda, club, etc.",
        ],
        "resultados": [
            "Módulo de inventario completamente funcional con CRUD de elementos, series y lotes.",
            "4 vistas principales implementadas: dashboard, lista de elementos, detalle de elemento, gestión de ubicaciones.",
            "Sistema de categorías con soporte visual de iconos.",
            "Buscador global integrado en el módulo de inventario.",
        ],
        "dificultades": [
            "El renderizado de tarjetas con valor 0 (cero) en React mostraba el número en pantalla en lugar de estar oculto. Se debió a la evaluación de expresiones condicionales con && que tratan el 0 como falsy pero lo renderizan.",
            "La invalidación de queries de React Query al eliminar categorías no funcionaba correctamente, requiriendo ajustes en las keys de caché.",
            "La integración del selector de ubicaciones en los formularios de series tuvo problemas con valores null/undefined que debieron manejarse.",
        ],
        "plan_siguiente": [
            "Implementar el módulo de elementos compuestos (plantillas).",
            "Comenzar el módulo de alquileres con cotizaciones.",
            "Implementar la reorganización del backend en módulos separados.",
        ],
    },
    {
        "numero": 6,
        "fecha": "13 de diciembre de 2025",
        "periodo": "13 de diciembre – 26 de diciembre de 2025",
        "horas": 35,
        "objetivo": "Implementar los elementos compuestos (plantillas de productos) y reorganizar el backend en módulos independientes.",
        "actividades": [
            "Elementos compuestos - Backend: Implementación del modelo de datos para elementos compuestos: una plantilla puede contener múltiples componentes del inventario. Soporte para grupos de componentes alternativos: por ejemplo, una carpa puede usar lona tipo A o tipo B como alternativa intercambiable. Endpoints CRUD para crear, listar, editar y eliminar plantillas compuestas. Consultas optimizadas para obtener plantillas con sus componentes y conteos.",
            "Elementos compuestos - Frontend: Formulario multi-paso para crear elementos compuestos: paso 1 (datos generales), paso 2 (selección de componentes), paso 3 (resumen). Vista de listado con tarjetas que muestran el conteo de componentes por plantilla. Modal de edición con las mismas funcionalidades del formulario de creación.",
            "Reorganización del backend: Reestructuración completa del backend en módulos separados: Inventario, Productos y Alquileres. Cada módulo contiene sus propios controllers, models y routes. Simplificación del archivo server.js de ~40 importaciones a una importación por módulo.",
            "Mejoras al módulo de inventario: Corrección del renderizado de 0 en tarjetas de ubicación. Nuevos tipos de ubicación para eventos (hotel, hacienda, club campestre). Corrección de la integración del selector de ubicaciones.",
        ],
        "resultados": [
            "Módulo de elementos compuestos funcional con soporte de alternativas.",
            "Backend reorganizado en 3 módulos principales con arquitectura limpia.",
            "12 PRs mergeados durante este periodo (#30 a #42).",
            "Correcciones de bugs críticos en el módulo de inventario.",
        ],
        "dificultades": [
            "La simplificación del modelo de datos para elementos compuestos requirió múltiples iteraciones. El diseño original era demasiado complejo y se simplificó para mantener la usabilidad.",
            "La reorganización del backend en módulos causó conflictos en las rutas de importación que debieron resolverse cuidadosamente.",
            "Se detectó un bug donde console.log de depuración quedaba en el código de producción del CategoriaModel.",
        ],
        "plan_siguiente": [
            "Implementar el módulo de alquileres: cotizaciones con vista previa PDF.",
            "Integrar tarifas de transporte por ciudad.",
            "Agregar sistema de calendario para visualización de eventos.",
        ],
    },
    {
        "numero": 7,
        "fecha": "27 de diciembre de 2025",
        "periodo": "27 de diciembre de 2025 – 9 de enero de 2026",
        "horas": 38,
        "objetivo": "Implementar el módulo de cotizaciones con vista previa estilo PDF, integrar tarifas de transporte y agregar el sistema de calendario.",
        "actividades": [
            "Módulo de cotizaciones: Implementación completa del flujo de cotizaciones: creación, edición, vista previa y generación PDF. Vista previa estilo PDF integrada en la interfaz web. Menú kebab (3 puntos) para acciones rápidas en cada cotización. Separación de hooks de cotizaciones en archivos individuales para mejor mantenibilidad.",
            "Tarifas de transporte: Tabla maestra de ciudades como entidad independiente. Integración de tarifas de transporte en el formulario de ciudades. Migración completa a ciudad_id como clave foránea. Eliminación del módulo redundante de tarifas que estaba duplicado.",
            "Sistema de calendario: Integración de FullCalendar con vistas de día, semana, mes y lista. Arquitectura modular del calendario para reutilización. Modal resumen con productos al hacer clic en un evento del calendario.",
            "Disponibilidad en tiempo real: Sistema de verificación de disponibilidad de productos por rango de fechas. Debounce en llamadas API para evitar múltiples peticiones simultáneas. Uso de cantidad de tabla elementos como fallback cuando no hay series.",
        ],
        "resultados": [
            "Módulo de cotizaciones completo con generación de PDF.",
            "Sistema de calendario integrado con FullCalendar.",
            "Verificación de disponibilidad funcionando en tiempo real.",
            "12 PRs mergeados (#47 a #59).",
        ],
        "dificultades": [
            "Las múltiples llamadas a la API de disponibilidad causaban problemas de rendimiento. Se implementó debounce y se deshabilitaron retries en mutaciones.",
            "La migración a ciudad_id requirió un script de migración cuidadoso para no perder datos existentes.",
            "La corrección del nombre de columna id_elemento en consultas de series generó errores que se detectaron tardíamente.",
            "La compatibilidad con MySQL safe mode requirió ajustes en las migraciones para evitar UPDATE sin WHERE.",
        ],
        "plan_siguiente": [
            "Implementar fechas de montaje/desmontaje en cotizaciones.",
            "Crear el módulo de operaciones con órdenes de trabajo.",
            "Agregar soporte de imágenes para elementos.",
        ],
    },
    {
        "numero": 8,
        "fecha": "10 de enero de 2026",
        "periodo": "10 de enero – 23 de enero de 2026",
        "horas": 32,
        "objetivo": "Implementar el módulo de operaciones con órdenes de trabajo, agregar soporte de imágenes y mejorar la gestión de cotizaciones.",
        "actividades": [
            "Módulo de operaciones: Implementación del flujo completo de operaciones: cargue → montaje → desmontaje → retorno → descargue. Órdenes de trabajo con asignación de empleados y vehículos. Historial de estados por orden de trabajo. Checklists de cargue y descargue para verificación de equipos.",
            "Soporte de imágenes: Upload de imágenes para elementos y productos de alquiler usando Multer. Visualización de fotos en modal durante el proceso de cargue para verificación visual. Almacenamiento de imágenes en el servidor con rutas relativas.",
            "Mejoras en cotizaciones: Fechas de montaje y desmontaje como campos independientes en la cotización. Contexto de alquiler visible en series y lotes (saber qué está alquilado y para qué evento). Corrección de módulos de edición de series y lotes.",
            "Auto-generación de lotes: Número de lote generado automáticamente basado en la fecha de creación. Rediseño de la página de detalle de elemento con stat cards como filtros interactivos. Unificación del estado 'nuevo' dentro de 'bueno' para simplificar el flujo.",
        ],
        "resultados": [
            "Módulo de operaciones funcional con flujo completo de 5 etapas.",
            "Soporte de imágenes implementado para elementos y productos.",
            "Auto-generación de números de lote.",
            "Rediseño de la vista de detalle de elemento.",
            "8 PRs mergeados (#60 a #67).",
        ],
        "dificultades": [
            "El envío del campo cantidad al crear elementos no se estaba realizando correctamente, causando que los elementos se crearan sin stock inicial.",
            "Los iconos SVG no se renderizaban correctamente en algunos componentes; se debió crear un componente IconoCategoria reutilizable.",
            "Las tablas inexistentes en el script de limpieza generaban errores que debieron manejarse con IF EXISTS.",
        ],
        "plan_siguiente": [
            "Implementar cronómetros en tiempo real para operaciones de montaje/desmontaje.",
            "Agregar gestión de depósitos en cotizaciones.",
            "Implementar historial de eventos por cliente.",
        ],
    },
    {
        "numero": 9,
        "fecha": "24 de enero de 2026",
        "periodo": "24 de enero – 6 de febrero de 2026",
        "horas": 34,
        "objetivo": "Implementar funcionalidades avanzadas: cronómetros en vivo, gestión de depósitos, historial de eventos por cliente y funcionalidad de repetir eventos.",
        "actividades": [
            "Cronómetros en tiempo real: Implementación de temporizadores en vivo para operaciones de montaje y desmontaje. Los cronómetros inician cuando se cambia el estado de la orden de trabajo y se detienen al finalizar. Modales de confirmación estilizados para cambios de estado. Auto-creación de la tabla orden_trabajo_historial_estados.",
            "Gestión de depósitos: Campo de valor de depósito agregado a las cotizaciones. Toggle de cobro de depósito (cobrar o no cobrar). Uso de valores del backend (resumen) en vez de recalcular subtotales en la vista previa. Corrección del error cobrar_deposito que fallaba por tipo de dato incorrecto.",
            "Historial de eventos por cliente: Nueva página de eventos completados con historial de productos alquilados. Filtro para mostrar solo eventos activos en la lista de cotizaciones. Visualización de productos alquilados y cantidades en el historial. Corrección de nombres de columnas en la consulta de historial de productos.",
            "Funcionalidad de repetir evento: Botón 'Repetir evento' que crea una nueva cotización con los mismos productos. Modal de formulario con campos de fecha resaltados para indicar que deben cambiarse. Conversión de valores DECIMAL de MySQL a números en JavaScript para evitar errores de tipo.",
        ],
        "resultados": [
            "Cronómetros en tiempo real funcionando para montaje y desmontaje.",
            "Sistema de depósitos integrado en cotizaciones.",
            "Historial completo de eventos por cliente con opción de repetir.",
            "Flujo de desmontaje ampliado con 2 nuevas etapas.",
            "6 PRs mergeados (#77 a #82).",
        ],
        "dificultades": [
            "Los hooks de contexto nunca se ejecutaban debido a una comparación estricta (===) con booleanos de MySQL, que devuelve 0 y 1 en lugar de true y false. Se solucionó convirtiendo explícitamente a booleano.",
            "El valor requiere_series sin convertir a booleano provocaba que se renderizara un 0 en pantalla.",
            "La consulta de checklist hacía referencia a una columna el.compuesto_id que no existía en la tabla, causando errores SQL.",
            "El stock en alquileres mostraba valores incorrectos porque se descontaba doblemente: una vez al crear la cotización y otra al confirmar el alquiler.",
        ],
        "plan_siguiente": [
            "Corregir el bug de stock cero en cotizaciones.",
            "Rediseñar la interfaz del módulo de inventario.",
            "Reorganizar la estructura completa del frontend y backend por módulos.",
        ],
    },
    {
        "numero": 10,
        "fecha": "7 de febrero de 2026",
        "periodo": "7 de febrero – 20 de febrero de 2026",
        "horas": 36,
        "objetivo": "Corregir bugs críticos de stock, rediseñar la interfaz de inventario y realizar la reorganización completa del proyecto por módulos.",
        "actividades": [
            "Corrección del bug de stock cero: Se identificó que las cotizaciones descontaban el stock al crearlas y nuevamente al confirmar el alquiler, provocando un doble descuento que dejaba productos con stock 0 cuando no debía. Se corrigió el mapeo de disponibilidad por producto en el selector. Pruebas exhaustivas del flujo completo: crear cotización → confirmar alquiler → verificar stock → devolver → verificar restauración de stock.",
            "Rediseño de la interfaz de inventario: Nueva interfaz del módulo de inventario con diseño más limpio y moderno. Tarjetas de productos con mejor visualización de stock y estados. Mejoras en la experiencia de usuario para navegación entre módulos.",
            "Reorganización completa del proyecto: Backend: Creación de 7 módulos independientes (auth, inventario, productos, alquileres, clientes, operaciones, configuracion), cada uno con su propia estructura de controllers/models/routes. Frontend: Reorganización de 247 archivos en 9 módulos + shared. Simplificación del servidor: server.js pasó de ~40 líneas de importación a una línea por módulo.",
            "Resolución de conflictos post-reorganización: Corrección de rutas de importación de ClienteModel y ConfiguracionModel en controladores que quedaron con rutas antiguas. Merge de conflictos entre la rama de reorganización y main.",
        ],
        "resultados": [
            "Bug crítico de doble descuento de stock corregido.",
            "Interfaz de inventario rediseñada con mejor UX.",
            "Proyecto completamente reorganizado en arquitectura modular: 7 módulos backend, 9 módulos frontend + shared, 261 endpoints API organizados por dominio.",
            "3 PRs mergeados (#83, #84, #85).",
        ],
        "dificultades": [
            "La reorganización de 247 archivos fue la tarea más compleja del proyecto. Cada archivo movido requería actualizar todas sus importaciones y las importaciones de los archivos que lo referenciaban.",
            "Algunos controladores seguían importando modelos con rutas relativas antiguas (../../models/) en lugar de las nuevas rutas modulares, causando errores en tiempo de ejecución.",
            "La resolución de conflictos de merge entre la rama de reorganización y main requirió cuidado para mantener los alias de importación consistentes.",
        ],
        "plan_siguiente": [
            "Pruebas de integración end-to-end.",
            "Documentación técnica final para el trabajo de grado.",
            "Despliegue en ambiente de producción.",
            "Capacitación al usuario final.",
        ],
    },
]

# Styles
header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
title_font = Font(name="Calibri", bold=True, size=14, color="2F5496")
subtitle_font = Font(name="Calibri", bold=True, size=11, color="2F5496")
normal_font = Font(name="Calibri", size=11)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
wrap_alignment = Alignment(wrap_text=True, vertical="top")
center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for b in bitacoras:
    ws = wb.create_sheet(title=f"Bitácora {b['numero']:02d}")

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 90

    row = 1
    # Title
    ws.merge_cells(f"A{row}:B{row}")
    cell = ws[f"A{row}"]
    cell.value = f"BITÁCORA DE TRABAJO DE GRADO - N° {b['numero']:02d}"
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center")
    row += 2

    # Info table
    info_fields = [
        ("Fecha", b["fecha"]),
        ("Periodo", b["periodo"]),
        ("Proyecto", "Sistema de Gestión de Inventario para Alquiler de Carpas"),
        ("Estudiante", "Anderson"),
        ("Horas dedicadas", f"{b['horas']} horas"),
    ]
    for label, value in info_fields:
        ws[f"A{row}"].value = label
        ws[f"A{row}"].font = Font(name="Calibri", bold=True, size=11)
        ws[f"A{row}"].border = thin_border
        ws[f"A{row}"].alignment = wrap_alignment
        ws[f"B{row}"].value = value
        ws[f"B{row}"].font = normal_font
        ws[f"B{row}"].border = thin_border
        ws[f"B{row}"].alignment = wrap_alignment
        row += 1

    row += 1

    # Sections
    sections = [
        ("OBJETIVO DEL PERIODO", [b["objetivo"]]),
        ("ACTIVIDADES REALIZADAS", b["actividades"]),
        ("RESULTADOS Y AVANCES", b["resultados"]),
        ("DIFICULTADES ENCONTRADAS", b["dificultades"]),
        ("PLAN PARA EL SIGUIENTE PERIODO", b["plan_siguiente"]),
    ]

    for section_title, items in sections:
        ws.merge_cells(f"A{row}:B{row}")
        cell = ws[f"A{row}"]
        cell.value = section_title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        ws[f"B{row}"].fill = header_fill
        row += 1

        for i, item in enumerate(items, 1):
            ws.merge_cells(f"A{row}:B{row}")
            cell = ws[f"A{row}"]
            if len(items) > 1:
                cell.value = f"{i}. {item}"
            else:
                cell.value = item
            cell.font = normal_font
            cell.alignment = wrap_alignment
            cell.border = thin_border
            ws[f"B{row}"].border = thin_border
            # Set row height based on content length
            ws.row_dimensions[row].height = max(30, (len(item) // 100 + 1) * 18)
            row += 1

        row += 1

# Remove default sheet
del wb["Sheet"]

# Save
output_path = "/home/user/aprendizaje-inventario-carpas/docs/bitacoras/Bitacoras_Trabajo_de_Grado.xlsx"
wb.save(output_path)
print(f"Excel guardado en: {output_path}")
